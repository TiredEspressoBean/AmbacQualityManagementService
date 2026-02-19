"""
Template generation for CSV/Excel import.

Generates downloadable templates with:
- Headers and field hints
- Sample data rows
- FK lookup sheets (Excel only)
- Instructions and field reference

Supports automatic model introspection for zero-config usage.
"""

import io
from typing import Any, Dict, List, Optional, Type

import pandas as pd
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


# Style constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HINT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HINT_FONT = Font(color="666666", italic=True, size=9)
REQUIRED_FONT = Font(color="C00000", bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Fields to skip in auto-generation (internal/system fields)
SKIP_FIELDS = {
    'id', 'tenant', 'created_at', 'updated_at', 'archived',
    'created_by', 'modified_by', 'version', 'previous_version',
    'is_current_version', 'classification',
}


class TemplateField:
    """
    Definition of a field for template generation.

    Attributes:
        name: Field/column name
        required: Whether field is required
        description: Help text for the field
        example: Example value to show in template
        choices: List of valid values (for enum fields)
        fk_model: Related model for FK lookup
        fk_display_field: Field to show in lookup (default: name)
        fk_value_field: Field to use as value (default: name or ERP_id)
    """

    def __init__(
        self,
        name: str,
        required: bool = False,
        description: str = "",
        example: Any = "",
        choices: Optional[List[str]] = None,
        fk_model: Optional[Type[models.Model]] = None,
        fk_display_field: str = "name",
        fk_value_field: Optional[str] = None,
    ):
        self.name = name
        self.required = required
        self.description = description
        self.example = example
        self.choices = choices
        self.fk_model = fk_model
        self.fk_display_field = fk_display_field
        self.fk_value_field = fk_value_field or fk_display_field

    @property
    def header_name(self) -> str:
        """Get header name with required marker."""
        if self.required:
            return f"{self.name}*"
        return self.name

    @property
    def hint_text(self) -> str:
        """Get hint text for the hint row."""
        parts = []
        if self.description:
            parts.append(self.description)
        if self.choices:
            parts.append(f"Values: {', '.join(self.choices)}")
        if self.fk_model:
            parts.append(f"See {self.fk_model.__name__} sheet")
        return " | ".join(parts) if parts else ""


class TemplateGenerator:
    """
    Generates import templates for a model.

    Usage:
        generator = TemplateGenerator(
            model_name="Parts",
            fields=[
                TemplateField("ERP_id", required=True, example="PRT-00001"),
                TemplateField("part_type", required=True, fk_model=PartTypes),
                TemplateField("status", choices=["pending", "in_progress", "completed"]),
            ]
        )

        # Generate Excel template with FK lookups
        xlsx_content = generator.generate_excel(tenant=request.user.tenant)

        # Generate simple CSV template
        csv_content = generator.generate_csv()
    """

    def __init__(
        self,
        model_name: str,
        fields: List[TemplateField],
        description: str = "",
    ):
        self.model_name = model_name
        self.fields = fields
        self.description = description

    def generate_csv(self) -> bytes:
        """
        Generate a simple CSV template.

        Returns header row and optional sample row.
        """
        output = io.StringIO()

        # Header row with required markers
        headers = [f.header_name for f in self.fields]
        output.write(",".join(headers) + "\n")

        # Sample row
        examples = [str(f.example) if f.example else "" for f in self.fields]
        output.write(",".join(examples) + "\n")

        return output.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility

    def generate_excel(
        self,
        tenant=None,
        include_lookups: bool = True,
        include_instructions: bool = True,
    ) -> bytes:
        """
        Generate an Excel template with multiple sheets.

        Args:
            tenant: Tenant for filtering FK lookup data
            include_lookups: Whether to include FK lookup sheets
            include_instructions: Whether to include instructions sheet

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Data sheet (main template)
        self._create_data_sheet(wb)

        # Instructions sheet
        if include_instructions:
            self._create_instructions_sheet(wb)

        # FK lookup sheets
        if include_lookups:
            self._create_lookup_sheets(wb, tenant)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_data_sheet(self, wb: Workbook):
        """Create the main data entry sheet."""
        ws = wb.active
        ws.title = "Data"

        # Row 1: Headers
        for col, field in enumerate(self.fields, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = field.header_name
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

            # Mark required fields
            if field.required:
                cell.font = Font(color="FFFFFF", bold=True)

            # Set column width based on content
            width = max(len(field.header_name), len(str(field.example or "")), 15)
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width + 2

        # Row 2: Hints
        for col, field in enumerate(self.fields, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = field.hint_text
            cell.fill = HINT_FILL
            cell.font = HINT_FONT
            cell.alignment = Alignment(wrap_text=True)

        # Row 3: Sample data
        for col, field in enumerate(self.fields, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = field.example
            cell.font = Font(italic=True, color="808080")

        # Add data validation for choice fields
        for col, field in enumerate(self.fields, start=1):
            if field.choices:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(field.choices)}"',
                    allow_blank=not field.required,
                )
                dv.error = f"Please select a valid {field.name}"
                dv.errorTitle = "Invalid Value"
                ws.add_data_validation(dv)
                # Apply to rows 4-1000 (data entry area)
                col_letter = ws.cell(row=1, column=col).column_letter
                dv.add(f"{col_letter}4:{col_letter}1000")

        # Freeze header rows
        ws.freeze_panes = "A3"

    def _create_instructions_sheet(self, wb: Workbook):
        """Create the instructions/field reference sheet."""
        ws = wb.create_sheet("Instructions")

        # Title
        ws.cell(row=1, column=1).value = f"{self.model_name} Import Template"
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)

        if self.description:
            ws.cell(row=2, column=1).value = self.description
            ws.cell(row=2, column=1).font = Font(italic=True)

        # Import modes
        ws.cell(row=4, column=1).value = "Import Modes:"
        ws.cell(row=4, column=1).font = Font(bold=True)
        ws.cell(row=5, column=1).value = "- create: Only create new records (error if exists)"
        ws.cell(row=6, column=1).value = "- update: Only update existing records (error if not found)"
        ws.cell(row=7, column=1).value = "- upsert (default): Create or update based on lookup fields"

        # Field reference header
        ws.cell(row=9, column=1).value = "Field Reference"
        ws.cell(row=9, column=1).font = Font(size=12, bold=True)

        # Field reference table headers
        headers = ["Field", "Required", "Description", "Valid Values", "Example"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Field reference rows
        for row, field in enumerate(self.fields, start=11):
            ws.cell(row=row, column=1).value = field.name
            ws.cell(row=row, column=1).font = Font(bold=field.required)

            ws.cell(row=row, column=2).value = "Yes" if field.required else "No"
            if field.required:
                ws.cell(row=row, column=2).font = REQUIRED_FONT

            ws.cell(row=row, column=3).value = field.description

            if field.choices:
                ws.cell(row=row, column=4).value = ", ".join(field.choices)
            elif field.fk_model:
                ws.cell(row=row, column=4).value = f"See '{field.fk_model.__name__}' sheet"

            ws.cell(row=row, column=5).value = str(field.example) if field.example else ""

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20

    def _create_lookup_sheets(self, wb: Workbook, tenant=None):
        """Create FK lookup sheets for related models."""
        # Get unique FK models
        fk_models = {}
        for field in self.fields:
            if field.fk_model and field.fk_model not in fk_models:
                fk_models[field.fk_model] = field

        for model, field in fk_models.items():
            self._create_lookup_sheet(wb, model, field, tenant)

    def _create_lookup_sheet(
        self,
        wb: Workbook,
        model: Type[models.Model],
        field: TemplateField,
        tenant=None,
    ):
        """Create a single FK lookup sheet."""
        ws = wb.create_sheet(model.__name__)

        # Get queryset
        qs = model.objects.all()

        # Apply tenant filter if model has tenant field
        if tenant and hasattr(model, 'tenant'):
            qs = qs.filter(tenant=tenant)

        # Filter out archived if applicable
        if hasattr(model, 'archived'):
            qs = qs.filter(archived=False)

        # Determine fields to include
        display_fields = [field.fk_value_field]
        if field.fk_display_field != field.fk_value_field:
            display_fields.append(field.fk_display_field)

        # Add ERP_id if available and not already included
        if hasattr(model, 'ERP_id') and 'ERP_id' not in display_fields:
            display_fields.append('ERP_id')

        # Add id for reference
        if 'id' not in display_fields:
            display_fields.append('id')

        # Build data
        data = []
        for obj in qs.order_by(field.fk_display_field)[:500]:  # Limit to 500 rows
            row = {}
            for f in display_fields:
                row[f] = getattr(obj, f, None)
            data.append(row)

        if not data:
            ws.cell(row=1, column=1).value = f"No {model.__name__} records found"
            return

        # Create DataFrame and write to sheet
        df = pd.DataFrame(data)

        # Write headers
        for col, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Write data
        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx).value = value

        # Auto-size columns
        for col in range(1, len(df.columns) + 1):
            max_length = max(
                len(str(df.columns[col-1])),
                df.iloc[:, col-1].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_length + 2, 50)


# ===== AUTOMATIC MODEL INTROSPECTION =====

def get_example_value(field: models.Field) -> Any:
    """Generate an example value based on field type."""
    if isinstance(field, models.CharField):
        if 'email' in field.name.lower():
            return "example@email.com"
        if 'name' in field.name.lower():
            return "Example Name"
        if 'id' in field.name.lower() or 'erp' in field.name.lower():
            return "EX-001"
        return "Example"
    elif isinstance(field, models.TextField):
        return "Description text"
    elif isinstance(field, models.IntegerField):
        return 100
    elif isinstance(field, (models.DecimalField, models.FloatField)):
        return 10.5
    elif isinstance(field, models.BooleanField):
        return "false"
    elif isinstance(field, (models.DateField, models.DateTimeField)):
        return "2024-12-31"
    elif isinstance(field, models.DurationField):
        return "01:30:00"
    return ""


def get_field_description(field: models.Field) -> str:
    """Generate description from field help_text or name."""
    if field.help_text:
        return str(field.help_text)
    # Convert field name to readable format
    name = field.name.replace('_', ' ').title()
    if isinstance(field, models.ForeignKey):
        return f"{name} (name, ID, or ERP_id)"
    return name


def introspect_model(model: Type[models.Model], skip_fields: Optional[set] = None) -> List[TemplateField]:
    """
    Automatically generate TemplateFields from a Django model.

    Introspects the model to determine:
    - Field names and types
    - Required vs optional fields
    - Choice fields
    - Foreign key relationships

    Args:
        model: Django model class
        skip_fields: Set of field names to skip (defaults to SKIP_FIELDS)

    Returns:
        List of TemplateField objects
    """
    skip = skip_fields or SKIP_FIELDS
    fields = []

    for model_field in model._meta.get_fields():
        # Skip reverse relations
        if model_field.auto_created and not model_field.concrete:
            continue

        # Skip specified fields
        if model_field.name in skip:
            continue

        # Skip non-editable fields
        if hasattr(model_field, 'editable') and not model_field.editable:
            continue

        # Handle ForeignKey
        if isinstance(model_field, models.ForeignKey):
            related_model = model_field.related_model

            # Determine best lookup field for the related model
            fk_display = 'name' if hasattr(related_model, 'name') else 'id'
            fk_value = fk_display

            fields.append(TemplateField(
                name=model_field.name,
                required=not model_field.null and not model_field.blank,
                description=get_field_description(model_field),
                example=f"Example {related_model.__name__}",
                fk_model=related_model,
                fk_display_field=fk_display,
                fk_value_field=fk_value,
            ))
            continue

        # Skip ManyToMany for now
        if isinstance(model_field, models.ManyToManyField):
            continue

        # Handle choice fields
        choices = None
        if hasattr(model_field, 'choices') and model_field.choices:
            choices = [str(c[0]) for c in model_field.choices]

        # Determine if required
        required = False
        if hasattr(model_field, 'blank') and hasattr(model_field, 'null'):
            required = not model_field.blank and not model_field.null
        elif hasattr(model_field, 'null'):
            required = not model_field.null

        fields.append(TemplateField(
            name=model_field.name,
            required=required,
            description=get_field_description(model_field),
            example=get_example_value(model_field),
            choices=choices,
        ))

    return fields


def template_from_model(
    model: Type[models.Model],
    description: Optional[str] = None,
    skip_fields: Optional[set] = None,
    field_overrides: Optional[Dict[str, Dict[str, Any]]] = None,
) -> TemplateGenerator:
    """
    Create a TemplateGenerator automatically from a Django model.

    This is the primary entry point for automatic template generation.

    Args:
        model: Django model class
        description: Optional description (auto-generated if not provided)
        skip_fields: Additional fields to skip
        field_overrides: Dict of field_name -> {attribute: value} to override auto-detected values

    Returns:
        TemplateGenerator instance

    Example:
        # Basic usage - fully automatic
        template = template_from_model(PartTypes)

        # With customization
        template = template_from_model(
            Parts,
            description="Import parts with part type lookup",
            field_overrides={
                'part_type': {'required': True, 'example': 'Widget'},
            }
        )
    """
    # Merge skip fields
    all_skip = SKIP_FIELDS.copy()
    if skip_fields:
        all_skip.update(skip_fields)

    # Introspect model
    fields = introspect_model(model, all_skip)

    # Apply overrides
    if field_overrides:
        for field in fields:
            if field.name in field_overrides:
                for attr, value in field_overrides[field.name].items():
                    setattr(field, attr, value)

    # Generate description if not provided
    if not description:
        description = f"Import {model.__name__} records. Use name or ID to reference related records."

    return TemplateGenerator(
        model_name=model.__name__,
        description=description,
        fields=fields,
    )


# Pre-built template configurations for common models

def get_part_types_template() -> TemplateGenerator:
    """Get template generator for PartTypes import."""
    return TemplateGenerator(
        model_name="PartTypes",
        description="Import part type definitions. ERP_id is used for matching existing records.",
        fields=[
            TemplateField(
                "name", required=True,
                description="Part type name",
                example="Widget Assembly"
            ),
            TemplateField(
                "ERP_id",
                description="External system ID for matching",
                example="PT-001"
            ),
            TemplateField(
                "ID_prefix",
                description="Prefix for auto-generated part IDs",
                example="WDG-"
            ),
            TemplateField(
                "itar_controlled",
                description="ITAR controlled item",
                example="false",
                choices=["true", "false"]
            ),
            TemplateField(
                "eccn",
                description="Export Control Classification Number",
                example="EAR99"
            ),
            TemplateField(
                "usml_category",
                description="USML Category if ITAR controlled",
                example="XI"
            ),
        ]
    )


def get_parts_template(part_types_model=None, orders_model=None) -> TemplateGenerator:
    """Get template generator for Parts import."""
    from Tracker.models import PartTypes, Orders

    return TemplateGenerator(
        model_name="Parts",
        description="Import individual parts. Use part_type name or ERP_id to link to part types.",
        fields=[
            TemplateField(
                "ERP_id",
                description="Part ID (auto-generated if blank)",
                example="PRT-00001"
            ),
            TemplateField(
                "part_type", required=True,
                description="Part type name or ERP_id",
                example="Widget Assembly",
                fk_model=PartTypes,
                fk_display_field="name",
                fk_value_field="name"
            ),
            TemplateField(
                "order",
                description="Order name or ID",
                example="ORD-2024-0001",
                fk_model=Orders,
                fk_display_field="name",
                fk_value_field="name"
            ),
            TemplateField(
                "part_status",
                description="Current status",
                example="PENDING",
                choices=["PENDING", "IN_PROGRESS", "AWAITING_QA", "COMPLETED", "QUARANTINED"]
            ),
        ]
    )


def get_orders_template() -> TemplateGenerator:
    """Get template generator for Orders import."""
    from Tracker.models import Companies

    return TemplateGenerator(
        model_name="Orders",
        description="Import customer orders. Company can be specified by name.",
        fields=[
            TemplateField(
                "name", required=True,
                description="Order name/reference",
                example="2024-ORD-001"
            ),
            TemplateField(
                "order_number",
                description="Auto-generated order number (optional override)",
                example=""
            ),
            TemplateField(
                "company",
                description="Company name",
                example="Acme Corp",
                fk_model=Companies,
                fk_display_field="name",
                fk_value_field="name"
            ),
            TemplateField(
                "customer_note",
                description="Notes visible to customer",
                example="Rush order - expedite"
            ),
            TemplateField(
                "order_status",
                description="Order status",
                example="PENDING",
                choices=["PENDING", "IN_PROGRESS", "SHIPPED", "DELIVERED", "CANCELLED"]
            ),
            TemplateField(
                "estimated_completion",
                description="Expected completion date (YYYY-MM-DD)",
                example="2024-12-31"
            ),
        ]
    )


def get_work_orders_template() -> TemplateGenerator:
    """Get template generator for WorkOrders import."""
    from Tracker.models import Orders, PartTypes

    return TemplateGenerator(
        model_name="WorkOrders",
        description="Import work orders. Link to orders by name and specify part type for process assignment.",
        fields=[
            TemplateField(
                "ERP_id", required=True,
                description="Work order ID",
                example="WO-001"
            ),
            TemplateField(
                "related_order",
                description="Order name or number",
                example="2024-ORD-001",
                fk_model=Orders,
                fk_display_field="name",
                fk_value_field="name"
            ),
            TemplateField(
                "part_type",
                description="Part type for process lookup",
                example="Widget Assembly",
                fk_model=PartTypes,
                fk_display_field="name",
                fk_value_field="name"
            ),
            TemplateField(
                "quantity", required=True,
                description="Number of parts to create",
                example="100"
            ),
            TemplateField(
                "workorder_status",
                description="Work order status",
                example="PENDING",
                choices=["PENDING", "IN_PROGRESS", "COMPLETED", "ON_HOLD", "CANCELLED"]
            ),
            TemplateField(
                "expected_completion",
                description="Expected completion date (YYYY-MM-DD)",
                example="2024-12-31"
            ),
            TemplateField(
                "notes",
                description="Internal notes",
                example="Priority build"
            ),
        ]
    )


def get_equipment_template() -> TemplateGenerator:
    """Get template generator for Equipment import."""
    from Tracker.models import EquipmentType

    return TemplateGenerator(
        model_name="Equipment",
        description="Import equipment records. Equipment type must exist or be created first.",
        fields=[
            TemplateField(
                "name", required=True,
                description="Equipment name",
                example="CNC Mill #1"
            ),
            TemplateField(
                "equipment_type", required=True,
                description="Equipment type name",
                example="CNC Machine",
                fk_model=EquipmentType,
                fk_display_field="name",
                fk_value_field="name"
            ),
            TemplateField(
                "serial_number",
                description="Manufacturer serial number",
                example="SN-12345"
            ),
            TemplateField(
                "manufacturer",
                description="Equipment manufacturer",
                example="Haas"
            ),
            TemplateField(
                "model_number",
                description="Model number",
                example="VF-2"
            ),
            TemplateField(
                "location",
                description="Physical location",
                example="Bay 3"
            ),
            TemplateField(
                "status",
                description="Equipment status",
                example="active",
                choices=["active", "maintenance", "retired"]
            ),
            TemplateField(
                "notes",
                description="Additional notes",
                example=""
            ),
        ]
    )
