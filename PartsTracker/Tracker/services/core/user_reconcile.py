"""
Bulk user reconcile — state-based "desired roster" semantics.

Each row in the bulk workbook describes the *desired* state of a user
identified by email:
  - If the email isn't in the tenant: create the user + send an invitation
    (the act of putting them in the workbook IS the invite).
  - If the email is in the tenant: update fields/groups/status to match
    the row.
  - Empty cells: no change to that field.

Re-running the same workbook is a no-op once state already matches.

Resend-invitation is intentionally NOT modeled here — it's a single-row
action on the User Management page, not a workbook concern.

This service is the single source of truth for what one row does. Both
the synchronous viewset path (<25 rows) and the Celery batch path
(>=25 rows) go through it.
"""
from __future__ import annotations

import io
from datetime import timedelta
from typing import Any, Dict, List, Optional, TYPE_CHECKING

from django.conf import settings
from django.db import transaction
from django.utils import timezone

if TYPE_CHECKING:
    from Tracker.models import Tenant, User


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

EMAIL_RE = r"^[^\s@]+@[^\s@]+\.[^\s@]+$"

# Map UI-friendly status labels to is_active. Case-insensitive on input.
_STATUS_TO_ACTIVE = {
    "active": True,
    "inactive": False,
}


def reconcile_user_row(*, row: Dict[str, Any], tenant: "Tenant", acting_user: "User") -> Dict[str, Any]:
    """Reconcile one row to the tenant's user roster.

    Args:
        row: dict with keys {email, first_name, last_name, group, status, message}.
            `groups` (plural) is also accepted as an alias for `group`.
            String values are stripped; empty strings treated as "no change".
        tenant: the tenant we're operating on.
        acting_user: the admin running the bulk action — recorded on
            UserInvitation.invited_by + audit fields where applicable.

    Returns:
        dict with one of these shapes:
          - {outcome: 'created', user_id, invitation_id, warnings: []}
          - {outcome: 'updated', user_id, changes: [field, ...], warnings: []}
          - {outcome: 'unchanged', user_id, warnings: []}
          - {outcome: 'error', error: str}

    Never raises — errors are returned in the dict so a bulk fanout can
    aggregate per-row results without try/except per call.
    """
    import re
    from django.core.exceptions import ValidationError
    from Tracker.models import User, TenantGroup

    # ---- Validate & normalize input ---------------------------------------
    email = str(row.get("email", "") or "").strip().lower()
    first_name = str(row.get("first_name", "") or "").strip()
    last_name = str(row.get("last_name", "") or "").strip()
    group_raw = str(row.get("group", row.get("groups", "")) or "").strip()
    status_raw = str(row.get("status", "") or "").strip()
    message = str(row.get("message", "") or "").strip()

    if not email:
        return {"outcome": "error", "error": "email is required"}
    if not re.match(EMAIL_RE, email):
        return {"outcome": "error", "error": f"email format invalid: {email!r}"}

    # `groups` accepts semicolon-separated names for multi-group rows.
    # Lone empty string after split → no groups specified (don't change).
    group_names: Optional[List[str]] = None
    if group_raw:
        group_names = [g.strip() for g in group_raw.split(";") if g.strip()]

    # Status → is_active. Blank cell = no change.
    desired_is_active: Optional[bool] = None
    if status_raw:
        normalized = status_raw.lower()
        if normalized not in _STATUS_TO_ACTIVE:
            return {
                "outcome": "error",
                "error": f"status must be 'Active' or 'Inactive', got {status_raw!r}",
            }
        desired_is_active = _STATUS_TO_ACTIVE[normalized]

    # ---- Resolve groups (validates they exist in this tenant) -------------
    target_groups: Optional[List[TenantGroup]] = None
    if group_names is not None:
        target_groups = []
        for name in group_names:
            tg = TenantGroup.objects.filter(tenant=tenant, name__iexact=name).first()
            if tg is None:
                return {
                    "outcome": "error",
                    "error": f"group {name!r} not found in tenant",
                }
            target_groups.append(tg)

    # ---- Lookup user ------------------------------------------------------
    existing = User.objects.filter(tenant=tenant, email__iexact=email).first()

    if existing is None:
        # CREATE + INVITE path
        return _create_and_invite(
            tenant=tenant,
            acting_user=acting_user,
            email=email,
            first_name=first_name,
            last_name=last_name,
            target_groups=target_groups or [],
            desired_is_active=desired_is_active,
            message=message,
        )

    # UPDATE path — bring existing user into line with the row
    return _reconcile_existing(
        user=existing,
        tenant=tenant,
        acting_user=acting_user,
        first_name=first_name,
        last_name=last_name,
        target_groups=target_groups,
        desired_is_active=desired_is_active,
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _create_and_invite(
    *,
    tenant: "Tenant",
    acting_user: "User",
    email: str,
    first_name: str,
    last_name: str,
    target_groups: List["TenantGroup"],
    desired_is_active: Optional[bool],
    message: str,
) -> Dict[str, Any]:
    """Create a new User in this tenant, assign groups, fire an invitation."""
    from django.db import IntegrityError
    from Tracker.models import User, UserInvitation, UserRole

    # New users default to Active unless the row explicitly says Inactive.
    is_active = True if desired_is_active is None else desired_is_active

    try:
        with transaction.atomic():
            # Username uniqueness: AbstractUser requires username; use email.
            user = User.objects.create(
                tenant=tenant,
                email=email,
                username=email,
                first_name=first_name,
                last_name=last_name,
                is_active=is_active,
            )
            user.set_unusable_password()  # invite flow sets password on accept
            user.save(update_fields=["password"])

            # Group memberships via UserRole through-table.
            for tg in target_groups:
                UserRole.objects.create(
                    user=user, group=tg, granted_by=acting_user,
                )

            invitation = UserInvitation.objects.create(
                user=user,
                invited_by=acting_user,
                token=UserInvitation.generate_token(),
                expires_at=timezone.now() + timedelta(days=7),
            )
    except IntegrityError as e:
        return {"outcome": "error", "error": f"DB integrity: {e}"}
    except Exception as e:  # noqa: BLE001 - surface to caller; never raise
        return {"outcome": "error", "error": f"create failed: {e}"}

    # Copyable signup link — onboarding must work even when email delivery is
    # off (matches UserInvitationSerializer.get_invitation_url / send_invitation).
    invitation_url = f"{settings.FRONTEND_URL}/signup?token={invitation.token}"

    # Send invitation email outside the transaction so a failed send doesn't
    # roll back the user creation. `immediate=False` queues via Celery.
    try:
        from Tracker.email_notifications import send_invitation_email
        send_invitation_email(invitation.id, immediate=False)
    except Exception as e:  # noqa: BLE001
        # User + invitation row are created; email dispatch failed. Surface
        # as a warning rather than an outcome=error so the caller knows the
        # account exists but the email may need a manual resend.
        return {
            "outcome": "created",
            "user_id": str(user.id),
            "invitation_id": str(invitation.id),
            "invitation_url": invitation_url,
            "warnings": [f"invitation row created but email dispatch failed: {e}"],
        }

    # `message` is captured on the invitation log but isn't currently
    # threaded into the email template — flag it for future wiring rather
    # than silently dropping the value.
    warnings: List[str] = []
    if message:
        warnings.append(
            "message field is not yet plumbed into the invite email template"
        )

    return {
        "outcome": "created",
        "user_id": str(user.id),
        "invitation_id": str(invitation.id),
        "invitation_url": invitation_url,
        "warnings": warnings,
    }


def _reconcile_existing(
    *,
    user: "User",
    tenant: "Tenant",
    acting_user: "User",  # noqa: ARG001 - reserved for audit fields, see below
    first_name: str,
    last_name: str,
    target_groups: Optional[List["TenantGroup"]],
    desired_is_active: Optional[bool],
) -> Dict[str, Any]:
    """Bring an existing User into line with the desired row state."""
    from Tracker.models import UserRole

    changes: List[str] = []
    update_fields: List[str] = []

    # Name fields — blank cell means "no change" (already handled by caller
    # passing "" only when the column was empty).
    if first_name and user.first_name != first_name:
        user.first_name = first_name
        update_fields.append("first_name")
        changes.append("first_name")
    if last_name and user.last_name != last_name:
        user.last_name = last_name
        update_fields.append("last_name")
        changes.append("last_name")

    # Status flip
    if desired_is_active is not None and user.is_active != desired_is_active:
        user.is_active = desired_is_active
        update_fields.append("is_active")
        changes.append("is_active")

    if update_fields:
        user.save(update_fields=update_fields)

    # Group sync — only if the column was populated.
    # We *replace* the user's current tenant-group membership with the
    # target set (idempotent). This matches the workbook's "describe the
    # desired state" framing.
    if target_groups is not None:
        target_ids = {tg.id for tg in target_groups}
        current_roles = list(user.user_roles.filter(group__tenant=tenant))
        current_ids = {r.group_id for r in current_roles}

        # Remove memberships not in target
        removed = current_ids - target_ids
        if removed:
            user.user_roles.filter(group_id__in=removed).delete()
            changes.append("groups_removed")
        # Add memberships in target but not current
        added = target_ids - current_ids
        if added:
            for tg in target_groups:
                if tg.id in added:
                    UserRole.objects.create(
                        user=user, group=tg, granted_by=acting_user,
                    )
            changes.append("groups_added")

    if not changes:
        return {"outcome": "unchanged", "user_id": str(user.id), "warnings": []}
    return {
        "outcome": "updated",
        "user_id": str(user.id),
        "changes": changes,
        "warnings": [],
    }


# ---------------------------------------------------------------------------
# Workbook template generation
# ---------------------------------------------------------------------------

# Workbook columns must match the keys consumed by `reconcile_user_row`.
# Order is the column order in the Data sheet.
TEMPLATE_COLUMNS: List[Dict[str, Any]] = [
    {"key": "email",      "label": "Email",      "required": True,  "hint": "Required. Lookup key — unknown emails get created + invited."},
    {"key": "first_name", "label": "First Name", "required": False, "hint": "Optional. Updates on existing users; populates new users on create."},
    {"key": "last_name",  "label": "Last Name",  "required": False, "hint": "Optional. Same as First Name."},
    {"key": "group",      "label": "Group",      "required": False, "hint": "Tenant group name. Multi-group: semicolon-separated (e.g. 'Operator;QA Inspector')."},
    {"key": "status",     "label": "Status",     "required": False, "hint": "Active or Inactive. Blank = no change."},
    {"key": "message",    "label": "Message",    "required": False, "hint": "Optional. Only used when creating new users (passed to invitation email)."},
]

# Statuses reference list. Hard-coded — these are the only valid values.
STATUS_REFERENCE = ["Active", "Inactive"]


def build_user_reconcile_template(tenant: "Tenant", *, populate: bool = False) -> bytes:
    """Return an .xlsx workbook for the bulk-reconcile flow.

    Structure:
      Sheet 0  Data          — admin-editable rows (the parsed sheet on import)
      Sheet 1  Instructions  — usage notes
      Sheet 2  Groups        — tenant-scoped group names (drives Group dropdown)
      Sheet 3  Statuses      — [Active, Inactive] (drives Status dropdown)

    Args:
        tenant: tenant whose group list seeds the Groups reference sheet.
        populate: if True, pre-fill the Data sheet with the tenant's current
            users (one row per user). This is the "snapshot-and-edit"
            workflow — analogous to the Parts CSV export. Admins edit the
            rows in Excel and re-upload to apply. If False (default), the
            Data sheet is empty (the "add new users" workflow).

    Excel data-validation dropdowns are wired via named ranges so they
    stay consistent if the admin extends the reference data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    from Tracker.models import TenantGroup, User

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    REQUIRED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    INSTRUCTIONS_FONT = Font(size=11)
    INSTRUCTIONS_HEADING = Font(size=14, bold=True)

    wb = Workbook()
    # Workbook() ships with an empty default sheet — drop it; we add named ones.
    default = wb.active
    wb.remove(default)

    # Important: Data MUST be the first sheet (index 0). The existing
    # `parse_excel_file` defaults to `sheet_name=0`, so any consumer that
    # uses the standard parse path needs the data on the first sheet.
    # Reference + instructions sheets follow after.
    #
    # We create them in this order: Data (placeholder), Groups, Statuses,
    # Instructions. Then populate Data last so the defined-name ranges for
    # the reference sheets already exist when validations are wired.

    # ---- Sheet 0: Data (placeholder — populated below) --------------------
    data_ws = wb.create_sheet("Data")

    # ---- Sheet 1: Instructions --------------------------------------------
    instr = wb.create_sheet("Instructions")
    instr.column_dimensions["A"].width = 110
    instructions_lines = [
        ("Bulk User Reconcile — Instructions", INSTRUCTIONS_HEADING),
        ("", None),
        ("This workbook describes the DESIRED state of users in this organization.", INSTRUCTIONS_FONT),
        ("Each row in the 'Data' sheet is reconciled by email:", INSTRUCTIONS_FONT),
        ("  • Email not in the org → the user is created and sent an invitation email.", INSTRUCTIONS_FONT),
        ("  • Email already in the org → fields, group membership, and status are updated to match the row.", INSTRUCTIONS_FONT),
        ("  • Empty cells → no change to that field.", INSTRUCTIONS_FONT),
        ("  • Re-running the same workbook is a no-op once state matches.", INSTRUCTIONS_FONT),
        ("", None),
        ("Resending invitations is NOT done from this workbook — use the per-row 'Resend invitation' action on the User Management page.", INSTRUCTIONS_FONT),
        ("", None),
        ("Columns:", Font(bold=True, size=12)),
    ]
    for col in TEMPLATE_COLUMNS:
        prefix = "  • " + col["label"] + (" (required)" if col["required"] else "")
        instructions_lines.append((f"{prefix}: {col['hint']}", INSTRUCTIONS_FONT))
    instructions_lines.extend([
        ("", None),
        ("Group cell can contain multiple groups separated by ';' — for example 'Operator;QA Inspector'.", INSTRUCTIONS_FONT),
        ("Status cell accepts only the values from the 'Statuses' reference sheet (Active or Inactive).", INSTRUCTIONS_FONT),
        ("", None),
        ("Tip: rows ≥25 are queued in the background and the page will poll for completion automatically.", Font(italic=True, color="666666")),
    ])
    for i, (text, font) in enumerate(instructions_lines, start=1):
        cell = instr.cell(row=i, column=1, value=text)
        if font is not None:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Sheet 2: Groups (reference) --------------------------------------
    # Build before Data sheet so we can wire data validation against it.
    groups_ws = wb.create_sheet("Groups")
    groups_ws.cell(row=1, column=1, value="Group Name").font = HEADER_FONT
    groups_ws.cell(row=1, column=1).fill = HEADER_FILL
    groups_ws.column_dimensions["A"].width = 30
    group_names = list(
        TenantGroup.objects.filter(tenant=tenant)
        .order_by("name")
        .values_list("name", flat=True)
    )
    for i, name in enumerate(group_names, start=2):
        groups_ws.cell(row=i, column=1, value=name)
    groups_max_row = max(2, len(group_names) + 1)

    # ---- Sheet 3: Statuses (reference) ------------------------------------
    statuses_ws = wb.create_sheet("Statuses")
    statuses_ws.cell(row=1, column=1, value="Status").font = HEADER_FONT
    statuses_ws.cell(row=1, column=1).fill = HEADER_FILL
    statuses_ws.column_dimensions["A"].width = 20
    for i, s in enumerate(STATUS_REFERENCE, start=2):
        statuses_ws.cell(row=i, column=1, value=s)
    statuses_max_row = len(STATUS_REFERENCE) + 1

    # Named ranges for data validation. Named ranges (rather than literal
    # cell refs) survive sheet renames and let admins extend the reference
    # tables without breaking the Data sheet's dropdowns.
    wb.defined_names["ref_groups"] = DefinedName(
        "ref_groups", attr_text=f"Groups!$A$2:$A${groups_max_row}",
    )
    wb.defined_names["ref_statuses"] = DefinedName(
        "ref_statuses", attr_text=f"Statuses!$A$2:$A${statuses_max_row}",
    )

    # ---- Populate the Data sheet (created at the top as sheet 0) ----------
    # Required-column indication is via the REQUIRED_FILL cell fill (yellow
    # tint on the header), NOT a " *" suffix on the label — that suffix
    # changes the normalized column name (`email_*` ≠ `email`) and breaks
    # round-trip parsing when the populated workbook is uploaded back.
    for col_idx, col in enumerate(TEMPLATE_COLUMNS, start=1):
        label = col["label"]
        cell = data_ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if col["required"] else HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        # Column widths sized for typical content.
        widths = {"email": 32, "first_name": 18, "last_name": 18, "group": 36, "status": 14, "message": 50}
        data_ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col["key"], 20)

    # ---- Optional: populate with current tenant users ---------------------
    if populate:
        # Map column key → column index for fast lookup while writing rows.
        col_idx_by_key = {col["key"]: i for i, col in enumerate(TEMPLATE_COLUMNS, start=1)}
        users_qs = (
            User.objects.filter(tenant=tenant)
            .order_by("email")
            .prefetch_related("user_roles__group")
        )
        for r, u in enumerate(users_qs, start=2):
            # Collect this user's tenant group names. `user_roles` is the
            # reverse name on UserRole.user; filter to TenantGroup
            # memberships in this tenant only (cross-tenant rows
            # shouldn't appear but we filter defensively).
            tenant_group_names = sorted({
                role.group.name
                for role in u.user_roles.all()
                if getattr(role.group, "tenant_id", None) == tenant.id
            })
            data_ws.cell(row=r, column=col_idx_by_key["email"], value=u.email)
            data_ws.cell(row=r, column=col_idx_by_key["first_name"], value=u.first_name or "")
            data_ws.cell(row=r, column=col_idx_by_key["last_name"], value=u.last_name or "")
            data_ws.cell(
                row=r,
                column=col_idx_by_key["group"],
                value=";".join(tenant_group_names) if tenant_group_names else "",
            )
            data_ws.cell(
                row=r,
                column=col_idx_by_key["status"],
                value="Active" if u.is_active else "Inactive",
            )
            # message column intentionally left blank — it's only meaningful
            # for new invitations, not for state captured from existing users.

    # Data validations on Group and Status columns. The validation range is
    # rows 2-200 so admins have room to paste without re-applying it.
    group_col = next(i for i, c in enumerate(TEMPLATE_COLUMNS, start=1) if c["key"] == "group")
    status_col = next(i for i, c in enumerate(TEMPLATE_COLUMNS, start=1) if c["key"] == "status")

    # Group dropdown — list-type, allow_blank=True so multi-group cells can
    # be typed manually as 'A;B'. The dropdown is a *suggestion* aid.
    group_dv = DataValidation(
        type="list",
        formula1="=ref_groups",
        allow_blank=True,
        showErrorMessage=False,  # don't block typed multi-values
        showDropDown=False,       # weird API: False means "show the dropdown arrow"
    )
    group_letter = get_column_letter(group_col)
    group_dv.add(f"{group_letter}2:{group_letter}200")
    data_ws.add_data_validation(group_dv)

    # Status dropdown — strict, only Active/Inactive accepted.
    status_dv = DataValidation(
        type="list",
        formula1="=ref_statuses",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid status",
        error="Status must be 'Active' or 'Inactive'.",
        showDropDown=False,
    )
    status_letter = get_column_letter(status_col)
    status_dv.add(f"{status_letter}2:{status_letter}200")
    data_ws.add_data_validation(status_dv)

    # Make the Data sheet the one the workbook opens to.
    wb.active = wb.sheetnames.index("Data")

    # Serialize to bytes.
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
