"""
Data Export Mixin for ViewSets.

Provides export endpoints for filtered data in CSV and Excel formats.
Auto-configures from model introspection when no custom config is provided.

Excel exports include:
- Main data sheet
- Reference sheets for foreign key lookups
- Data validation dropdowns for FK fields
- Conditional formatting for required fields
- Instructions sheet with field documentation
"""

import io
import re
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
from django.db import models
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, OpenApiParameter
from rest_framework.decorators import action

# openpyxl imports for advanced Excel features
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# Fields to skip in auto-export
SKIP_EXPORT_FIELDS = {
    'tenant', 'created_by', 'modified_by', 'classification',
    'previous_version', 'is_current_version',
}

# Fields that are typically required
COMMON_REQUIRED_FIELDS = {'id', 'name', 'ERP_id'}


def get_exportable_fields(model) -> List[str]:
    """
    Get list of exportable fields from a model via introspection.

    Includes regular fields and common FK display fields (e.g., part_type__name).
    """
    fields = []

    for field in model._meta.get_fields():
        # Skip reverse relations
        if field.auto_created and not field.concrete:
            continue

        # Skip system fields
        if field.name in SKIP_EXPORT_FIELDS:
            continue

        # Handle ForeignKey - add both ID and name lookup
        if isinstance(field, models.ForeignKey):
            fields.append(field.name)
            # Add __name lookup if related model has name field
            related_model = field.related_model
            if hasattr(related_model, 'name'):
                fields.append(f'{field.name}__name')
            continue

        # Skip ManyToMany
        if isinstance(field, models.ManyToManyField):
            continue

        # Include regular fields
        if hasattr(field, 'name'):
            fields.append(field.name)

    return fields


def get_fk_fields(model) -> Dict[str, models.ForeignKey]:
    """
    Get all ForeignKey fields from a model.

    Returns dict mapping field name to the field object.
    """
    fk_fields = {}

    for field in model._meta.get_fields():
        if isinstance(field, models.ForeignKey):
            if field.name not in SKIP_EXPORT_FIELDS:
                fk_fields[field.name] = field

    return fk_fields


def get_choice_fields(model) -> Dict[str, List[str]]:
    """
    Get all fields with choices (enums) from a model.

    Returns dict mapping field name to list of valid choice values.
    """
    choice_fields = {}

    for field in model._meta.get_fields():
        if field.auto_created and not field.concrete:
            continue
        if field.name in SKIP_EXPORT_FIELDS:
            continue

        # Check for choices
        if hasattr(field, 'choices') and field.choices:
            # Handle both tuples and enum-style choices
            choices = []
            for choice in field.choices:
                if isinstance(choice, (list, tuple)) and len(choice) >= 1:
                    choices.append(str(choice[0]))
                else:
                    choices.append(str(choice))
            if choices:
                choice_fields[field.name] = choices

    return choice_fields


def get_boolean_fields(model) -> List[str]:
    """
    Get all boolean fields from a model.

    Returns list of field names.
    """
    boolean_fields = []

    for field in model._meta.get_fields():
        if field.auto_created and not field.concrete:
            continue
        if field.name in SKIP_EXPORT_FIELDS:
            continue

        if isinstance(field, (models.BooleanField, models.NullBooleanField)):
            boolean_fields.append(field.name)

    return boolean_fields


def get_field_info(model) -> Dict[str, Dict[str, Any]]:
    """
    Get detailed field information for documentation.

    Returns dict with field metadata including type, required, choices, etc.
    """
    field_info = {}

    for field in model._meta.get_fields():
        if field.auto_created and not field.concrete:
            continue
        if field.name in SKIP_EXPORT_FIELDS:
            continue
        if isinstance(field, models.ManyToManyField):
            continue

        info = {
            'name': field.name,
            'type': type(field).__name__,
            'required': not getattr(field, 'blank', True) or not getattr(field, 'null', True),
            'description': getattr(field, 'help_text', '') or '',
            'choices': None,
            'related_model': None,
        }

        # Get choices if available
        if hasattr(field, 'choices') and field.choices:
            info['choices'] = [str(c[0]) for c in field.choices]

        # Get related model for FKs
        if isinstance(field, models.ForeignKey):
            info['related_model'] = field.related_model.__name__
            info['description'] = info['description'] or f"Reference to {field.related_model.__name__}"

        field_info[field.name] = info

    return field_info


def get_field_label(field_name: str) -> str:
    """Convert field name to human-readable label."""
    # Handle __ lookups
    if '__' in field_name:
        parts = field_name.split('__')
        return ' '.join(p.replace('_', ' ').title() for p in parts)

    return field_name.replace('_', ' ').title()


def sanitize_sheet_name(name: str) -> str:
    """Sanitize a string to be a valid Excel sheet name."""
    # Excel sheet names: max 31 chars, no []:*?/\
    name = re.sub(r'[\[\]:*?/\\]', '', name)
    return name[:31]


def make_excel_safe_name(name: str) -> str:
    """Convert a string to a valid Excel named range identifier."""
    # Must start with letter or underscore, only alphanumeric and underscore
    safe = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    if safe and safe[0].isdigit():
        safe = '_' + safe
    return safe


class DataExportMixin:
    """
    Mixin to add CSV/Excel export functionality to ViewSets.

    Adds endpoints:
    - GET /export/csv/ - Download filtered data as CSV
    - GET /export/xlsx/ - Download filtered data as Excel (with reference sheets)

    Excel exports include:
    - Main data sheet with your filtered data
    - Reference sheets for each foreign key relationship
    - Data validation dropdowns for FK columns
    - Conditional formatting highlighting required fields
    - Instructions sheet documenting all fields

    Works automatically with any model through introspection.
    Override attributes for customization:

    - export_fields: List of fields to export (auto-detected if not set)
    - export_filename: Base filename for exports (model name if not set)
    - export_field_labels: Dict mapping field names to display labels
    - export_include_references: Whether to include FK reference sheets (default True)

    Example:
        # Basic - just add the mixin, everything auto-configured
        class PartsViewSet(DataExportMixin, viewsets.ModelViewSet):
            queryset = Parts.objects.all()

        # Custom - specify fields and labels
        class PartsViewSet(DataExportMixin, viewsets.ModelViewSet):
            export_fields = ['ERP_id', 'part_type__name', 'part_status']
            export_field_labels = {'ERP_id': 'Part ID', 'part_type__name': 'Part Type'}
    """

    export_fields: Optional[List[str]] = None
    export_filename: Optional[str] = None
    export_field_labels: Dict[str, str] = {}
    export_include_references: bool = True

    def _get_model(self):
        """Get the model class from queryset."""
        if hasattr(self, 'queryset') and self.queryset is not None:
            return self.queryset.model
        return None

    def get_export_fields(self) -> List[str]:
        """
        Get the list of fields to export.

        Priority:
        1. Query param ?fields=id,name,status (user override)
        2. self.export_fields (class attribute)
        3. Auto-detected from model (all non-system fields)
        """
        # Check query params first
        if hasattr(self, 'request'):
            fields_param = self.request.query_params.get('fields')
            if fields_param:
                return [f.strip() for f in fields_param.split(',')]

        # Use class attribute if specified
        if self.export_fields:
            return self.export_fields

        # Auto-detect from model
        model = self._get_model()
        if model:
            return get_exportable_fields(model)

        return ['id']

    def get_export_filename(self, format: str) -> str:
        """
        Get the filename for the export.

        Priority:
        1. Query param ?filename=custom
        2. self.export_filename (class attribute)
        3. Model name
        """
        if hasattr(self, 'request'):
            filename_param = self.request.query_params.get('filename')
            if filename_param:
                base_name = filename_param.rsplit('.', 1)[0]
                return f"{base_name}.{format}"

        if self.export_filename:
            return f"{self.export_filename}.{format}"

        model = self._get_model()
        if model:
            return f"{model.__name__.lower()}_export.{format}"

        return f"export.{format}"

    def get_export_field_labels(self) -> Dict[str, str]:
        """
        Get field label mapping.

        Merges auto-generated labels with custom labels.
        """
        labels = {}

        # Auto-generate labels for all fields
        for field in self.get_export_fields():
            labels[field] = get_field_label(field)

        # Override with custom labels
        labels.update(self.export_field_labels)

        return labels

    def get_export_queryset(self):
        """
        Get the queryset for export.

        Uses the same filtering as the list view.
        """
        return self.filter_queryset(self.get_queryset())

    def _convert_df_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert DataFrame types for Excel compatibility."""
        for col in df.columns:
            # Handle timezone-aware datetimes (openpyxl doesn't support them)
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:
                    if df[col].dt.tz is not None:
                        df[col] = df[col].dt.tz_localize(None)
                except (AttributeError, TypeError):
                    pass
            # Convert object columns (UUIDs, complex objects, etc.)
            elif df[col].dtype == 'object':
                df[col] = df[col].apply(
                    lambda x: str(x) if x is not None and not isinstance(x, (str, int, float, bool)) else x
                )
        return df

    def prepare_export_data(self, queryset, fields: List[str], apply_labels: bool = True) -> pd.DataFrame:
        """
        Prepare data for export as a DataFrame.

        Handles related field lookups (field__subfield) and applies labels.
        """
        # Separate simple fields from related fields
        simple_fields = []
        related_lookups = {}

        for field in fields:
            if '__' in field:
                # Related field lookup
                parts = field.split('__')
                related_lookups[field] = parts
            else:
                simple_fields.append(field)

        # Get data using values()
        if simple_fields:
            data = list(queryset.values(*simple_fields))
        else:
            data = list(queryset.values('id'))

        # Add related field values
        if related_lookups:
            # Re-fetch with related objects for lookups
            for i, obj in enumerate(queryset):
                for field_key, parts in related_lookups.items():
                    value = obj
                    for part in parts:
                        value = getattr(value, part, None)
                        if value is None:
                            break
                    if i < len(data):
                        data[i][field_key] = value

        # Create DataFrame
        df = pd.DataFrame(data)

        # Convert types for Excel compatibility
        df = self._convert_df_for_excel(df)

        # Reorder columns to match requested fields order
        existing_cols = [f for f in fields if f in df.columns]
        if existing_cols:
            df = df[existing_cols]

        # Apply field labels for column names
        if apply_labels:
            labels = self.get_export_field_labels()
            rename_map = {f: labels.get(f, f) for f in df.columns if f in labels}
            if rename_map:
                df.rename(columns=rename_map, inplace=True)

        return df

    def _get_reference_data(self, fk_field: models.ForeignKey, limit: int = 1000) -> pd.DataFrame:
        """
        Get reference data for a foreign key field.

        Returns DataFrame with id and display columns for the related model.
        """
        related_model = fk_field.related_model

        # Determine display field (prefer name, then fall back to str)
        if hasattr(related_model, 'name'):
            display_field = 'name'
        else:
            display_field = None

        # Build queryset - respect tenant if applicable
        qs = related_model.objects.all()

        # Try to filter by tenant if the model has tenant field
        if hasattr(self, 'request') and hasattr(self.request, 'tenant'):
            if hasattr(related_model, 'tenant'):
                qs = qs.filter(tenant=self.request.tenant)

        # Limit results
        qs = qs[:limit]

        # Get data
        if display_field:
            data = list(qs.values('id', display_field))
            df = pd.DataFrame(data)
            if not df.empty:
                df.columns = ['id', 'name']
        else:
            data = [{'id': str(obj.pk), 'name': str(obj)} for obj in qs]
            df = pd.DataFrame(data)

        # Convert types
        df = self._convert_df_for_excel(df)

        return df

    def _create_instructions_sheet(self, ws, model, field_info: Dict[str, Dict], fk_fields: Dict[str, models.ForeignKey]):
        """Create the instructions sheet with field documentation."""
        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF')
        required_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        # Title
        ws['A1'] = f'{model.__name__} Import/Export Guide'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        # Headers
        headers = ['Field Name', 'Required', 'Type', 'Description', 'Valid Values / Reference']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Field rows
        row = 4
        for field_name, info in field_info.items():
            ws.cell(row=row, column=1, value=field_name)

            required_cell = ws.cell(row=row, column=2, value='Yes' if info['required'] else 'No')
            if info['required']:
                required_cell.fill = required_fill

            ws.cell(row=row, column=3, value=info['type'])
            ws.cell(row=row, column=4, value=info['description'])

            # Valid values
            if info['choices']:
                ws.cell(row=row, column=5, value=', '.join(info['choices']))
            elif info['related_model']:
                ws.cell(row=row, column=5, value=f"See '{info['related_model']}' sheet")

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30

    def _create_reference_sheet(self, wb: Workbook, fk_name: str, fk_field: models.ForeignKey) -> Optional[Dict[str, Any]]:
        """
        Create a reference sheet for a foreign key.

        Returns dict with sheet info if successful:
        - range_name: Named range for dropdown
        - sheet_name: The worksheet name
        - max_row: Last row of data
        """
        ref_df = self._get_reference_data(fk_field)

        if ref_df.empty:
            return None

        # Create sheet
        related_name = fk_field.related_model.__name__
        sheet_name = sanitize_sheet_name(related_name)

        # Handle duplicate sheet names
        existing_names = [ws.title for ws in wb.worksheets]
        if sheet_name in existing_names:
            sheet_name = sanitize_sheet_name(f"{fk_name}_{related_name}")[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Style settings
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Write headers
        ws.cell(row=1, column=1, value='ID').font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value='Name').font = header_font
        ws.cell(row=1, column=2).fill = header_fill

        # Write data
        for row_idx, row_data in enumerate(ref_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        ws.column_dimensions['A'].width = 40  # UUID width
        ws.column_dimensions['B'].width = 30

        # Create named range for the name column (for dropdowns)
        range_name = make_excel_safe_name(f"ref_{fk_name}")
        max_row = len(ref_df) + 1

        # Named range for the name column (column B, rows 2 to end)
        from openpyxl.workbook.defined_name import DefinedName
        ref = f"'{sheet_name}'!$B$2:$B${max_row}"
        defn = DefinedName(range_name, attr_text=ref)
        wb.defined_names[range_name] = defn

        return {
            'range_name': range_name,
            'sheet_name': sheet_name,
            'max_row': max_row,
        }

    def _add_data_validation(self, ws, col_idx: int, range_name: str, max_row: int):
        """Add dropdown data validation to a column."""
        col_letter = get_column_letter(col_idx)

        # Create data validation with reference to named range
        dv = DataValidation(
            type='list',
            formula1=f'={range_name}',
            allow_blank=True,
            showDropDown=False,  # False means show dropdown (confusing API)
            showErrorMessage=True,
            errorTitle='Invalid Value',
            error='Please select a value from the dropdown list.',
        )

        # Apply to column (rows 2 to max_row, skipping header)
        dv.add(f'{col_letter}2:{col_letter}{max_row}')
        ws.add_data_validation(dv)

    def _add_inline_data_validation(self, ws, col_idx: int, values: List[str], max_row: int):
        """Add dropdown data validation with inline values (for small lists like enums)."""
        col_letter = get_column_letter(col_idx)

        # Excel inline list format: "value1,value2,value3"
        # Max length is ~255 chars, so this works for small choice lists
        values_str = ','.join(str(v) for v in values)

        dv = DataValidation(
            type='list',
            formula1=f'"{values_str}"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle='Invalid Value',
            error='Please select a value from the dropdown list.',
        )

        dv.add(f'{col_letter}2:{col_letter}{max_row}')
        ws.add_data_validation(dv)

    def _create_excel_export(self, queryset, fields: List[str], include_references: bool = True) -> bytes:
        """
        Create a full-featured Excel export with reference sheets.

        Returns the Excel file as bytes.
        """
        model = self._get_model()
        fk_fields = get_fk_fields(model) if model else {}
        choice_fields = get_choice_fields(model) if model else {}
        boolean_fields = get_boolean_fields(model) if model else []
        field_info = get_field_info(model) if model else {}

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active

        # Create Instructions sheet first
        instructions_ws = wb.create_sheet(title='Instructions', index=0)
        if model:
            self._create_instructions_sheet(instructions_ws, model, field_info, fk_fields)

        # Create reference sheets and track info for formulas
        ref_sheet_info = {}  # fk_field_name -> {range_name, sheet_name, max_row}
        if include_references:
            for fk_name, fk_field in fk_fields.items():
                info = self._create_reference_sheet(wb, fk_name, fk_field)
                if info:
                    ref_sheet_info[fk_name] = info

        # Create main data sheet
        data_ws = wb.create_sheet(title='Data', index=1)

        # Prepare data (without labels for now, we'll add custom headers)
        df = self.prepare_export_data(queryset, fields, apply_labels=False)

        # Style settings
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        required_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        formula_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Light green for formula columns

        # Track columns for data validation and formulas
        fk_name_columns = {}  # col_idx -> fk_name (for name columns with dropdowns)
        fk_id_columns = {}    # col_idx -> fk_name (for ID columns that need formulas)
        choice_columns = {}   # col_idx -> list of choices (for enum dropdowns)
        boolean_columns = []  # col_idx (for True/False dropdowns)

        # Write headers with formatting
        labels = self.get_export_field_labels()
        for col_idx, field_name in enumerate(df.columns, 1):
            label = labels.get(field_name, get_field_label(field_name))

            # Mark required fields
            is_required = field_info.get(field_name, {}).get('required', False)
            if is_required:
                label = f'{label}*'

            # Check if this is an FK ID column (will have formula)
            is_fk_id = field_name in fk_fields and field_name in ref_sheet_info
            if is_fk_id:
                label = f'{label} (auto)'  # Mark as auto-calculated

            cell = data_ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

            # Track FK columns
            base_field = field_name.split('__')[0]
            if base_field in fk_fields and field_name.endswith('__name'):
                fk_name_columns[col_idx] = base_field
            elif field_name in fk_fields and field_name in ref_sheet_info:
                fk_id_columns[col_idx] = field_name

            # Track choice/enum columns
            if field_name in choice_fields:
                choice_columns[col_idx] = choice_fields[field_name]

            # Track boolean columns
            if field_name in boolean_fields:
                boolean_columns.append(col_idx)

        # Build mapping of FK name to name column letter (for INDEX/MATCH formulas)
        fk_name_col_letters = {}  # fk_name -> column letter
        for col_idx, fk_name in fk_name_columns.items():
            fk_name_col_letters[fk_name] = get_column_letter(col_idx)

        # Write data rows
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                field_name = df.columns[col_idx - 1]

                # Check if this is an FK ID column that should have a formula
                if col_idx in fk_id_columns:
                    fk_name = fk_id_columns[col_idx]
                    if fk_name in ref_sheet_info and fk_name in fk_name_col_letters:
                        info = ref_sheet_info[fk_name]
                        name_col = fk_name_col_letters[fk_name]
                        sheet = info['sheet_name']
                        max_ref_row = info['max_row']

                        # INDEX/MATCH formula: look up ID based on name selection
                        # =IF(NameCol="","",INDEX(Sheet!$A$2:$A$max,MATCH(NameCol,Sheet!$B$2:$B$max,0)))
                        formula = (
                            f'=IF({name_col}{row_idx}="","",'
                            f"INDEX('{sheet}'!$A$2:$A${max_ref_row},"
                            f"MATCH({name_col}{row_idx},'{sheet}'!$B$2:$B${max_ref_row},0)))"
                        )
                        cell = data_ws.cell(row=row_idx, column=col_idx, value=formula)
                        cell.fill = formula_fill  # Light green to indicate formula
                        continue

                cell = data_ws.cell(row=row_idx, column=col_idx, value=value)

                # Highlight required field cells that are empty
                is_required = field_info.get(field_name, {}).get('required', False)
                if is_required and (value is None or value == ''):
                    cell.fill = required_fill

        # Add data validation for FK name columns
        max_row = max(len(df) + 1, 100)  # At least 100 rows for new entries
        for col_idx, fk_name in fk_name_columns.items():
            if fk_name in ref_sheet_info:
                self._add_data_validation(data_ws, col_idx, ref_sheet_info[fk_name]['range_name'], max_row)

        # Add data validation for choice/enum columns
        for col_idx, choices in choice_columns.items():
            self._add_inline_data_validation(data_ws, col_idx, choices, max_row)

        # Add data validation for boolean columns
        for col_idx in boolean_columns:
            self._add_inline_data_validation(data_ws, col_idx, ['True', 'False'], max_row)

        # Auto-adjust column widths
        for col_idx, field_name in enumerate(df.columns, 1):
            max_length = len(labels.get(field_name, field_name)) + 2
            # Sample first 100 rows for width
            for row_idx in range(2, min(len(df) + 2, 102)):
                cell_value = data_ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            data_ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # Freeze header row
        data_ws.freeze_panes = 'A2'

        # Remove the default empty sheet if it still exists
        if default_sheet.title == 'Sheet' and default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @extend_schema(
        parameters=[
            OpenApiParameter(
                name='fields',
                description='Comma-separated list of fields to export',
                required=False,
                type=str,
            ),
            OpenApiParameter(
                name='filename',
                description='Custom filename for the download',
                required=False,
                type=str,
            ),
            OpenApiParameter(
                name='include_references',
                description='Include FK reference sheets in Excel export (default: true)',
                required=False,
                type=bool,
            ),
        ],
        responses={
            200: {
                'type': 'string',
                'format': 'binary',
                'description': 'File download'
            }
        },
        description='Export filtered data to CSV or Excel format.',
        tags=['Import/Export'],
    )
    @action(detail=False, methods=['get'], url_path=r'export/(?P<export_format>csv|xlsx)')
    def export_data(self, request, export_format: str):
        """
        Export filtered data to CSV or Excel.

        URL path determines format:
        - /export/csv/ - CSV format (simple, just data)
        - /export/xlsx/ - Excel format (includes reference sheets, validation, formatting)

        Query params:
        - fields: Comma-separated field names
        - filename: Custom filename
        - include_references: Include FK reference sheets (xlsx only, default true)

        Respects all filters, search, and ordering applied to the list view.
        """
        # Get filtered queryset
        queryset = self.get_export_queryset()

        # Get fields to export
        fields = self.get_export_fields()

        # Generate filename
        filename = self.get_export_filename(export_format)

        # Create response
        if export_format == 'csv':
            # Simple CSV export
            df = self.prepare_export_data(queryset, fields)
            output = io.StringIO()
            df.to_csv(output, index=False)
            content = output.getvalue().encode('utf-8-sig')
            content_type = 'text/csv'
        else:
            # Full-featured Excel export
            include_refs = request.query_params.get('include_references', 'true').lower() != 'false'
            content = self._create_excel_export(queryset, fields, include_references=include_refs)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
